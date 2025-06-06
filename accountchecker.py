from modules.AccountChecker import AccountChecker
import os
from openpyxl import load_workbook

def main():
    DATA_FILE = "data.xlsx"
    BACKUP_FILE = "backup.xlsx"
    
    if os.path.exists(DATA_FILE):
        workbook = load_workbook(DATA_FILE)
        sheet = workbook.active

        # Tạo file backup trước khi chỉnh sửa
        workbook.save(BACKUP_FILE)
        print(f"🔒 Đã tạo file backup: {BACKUP_FILE}")
    
        bot = AccountChecker()

        # Lưu các chỉ số dòng cần xóa
        rows_to_delete = []
        # Duyệt ngược để xóa an toàn, bỏ qua dòng tiêu đề (giả sử tiêu đề ở dòng 1)
        for row_idx in range(sheet.max_row, 1, -1):  # từ dưới lên, bỏ qua header
            value = sheet.cell(row=row_idx, column=2).value
            print(value)
            result = bot.run(value)
            if result is False:
                rows_to_delete.append(row_idx)
        
        # Xóa từng dòng
        for idx in rows_to_delete:
            sheet.delete_rows(idx)

        # Lưu lại file
        workbook.save(DATA_FILE)
        print(f"✅ Đã cập nhật và lưu lại file: {DATA_FILE}")

if __name__ == "__main__":
    main()
