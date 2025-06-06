import requests
import json
from modules.TempMailClient import TempMailClient
from faker import Faker
import re
import string
import random
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from threading import Lock
import time

class Bot:
    def __init__(self, token, headless_mode=False, chromedriver: str="driver/chromedriver.exe", wait_sec=5):
        self.client = TempMailClient(token)
        self.state_drop = None
        self.driver = None
        self._driver_lock = Lock()
        self.headless_mode = headless_mode
        self.chromedriver = chromedriver
        self.wait_sec = wait_sec
        
    def _setup_driver(self):
        opts = webdriver.ChromeOptions()
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_argument('--log-level=3')
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_argument("--start-maximized")
        opts.add_experimental_option("prefs", {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False
        })
        if self.headless_mode:
            opts.add_argument('--headless')
        self.driver = webdriver.Chrome(service=Service(str(self.chromedriver)), options=opts)
        self.driver.get("https://ss2.sfcollege.edu/sr/AdmissionApplication/#/citizenship#top")
        
    def get_captcha_token(self):
        with self._driver_lock:
            if not self.driver:
                self._setup_driver()
            captchaToken = self.driver.execute_async_script('''
            var done = arguments[0];
            grecaptcha.ready(function() {
                grecaptcha.execute("6Letiu8UAAAAABdGKo97pgXWLi3sjHQLFDhyErZi", {action: 'passwordSecurity'}).then(function(token) {
                    done(token);
                }, function errorCallback(response) {
                    done("None")
                });
            });
            ''')
            print(captchaToken)
            return captchaToken if captchaToken and captchaToken != "None" else None

    def fake_profile(self):
        f = Faker()
        dob = f.date_of_birth(minimum_age=18, maximum_age=30)
        return {
            "first_name": f.first_name(),
            "last_name":  f.last_name(),
            "gender":     f.random_element(["Male", "Female"]),
            "birthdate":  dob,
            "ssn":        f.ssn(),
            "street":     f.street_address(),
            "city":       f.city(),
            "state":      f.state(),
            "zip":        f.zipcode()
        }

    def extract_code(self, html):
        m = re.search(r'<strong>(\d+)</strong>', html)
        return m.group(1) if m else ""

    def generate_username(self, length=9):
        chars = string.ascii_letters + string.digits
        return "".join(random.choices(chars, k=length))

    def generate_password(self, username="", length=12):
        if length < 8:
            raise ValueError("Password must be ≥ 8 chars")
        pools = [
            random.choice(string.ascii_uppercase),
            random.choice(string.ascii_lowercase),
            random.choice(string.digits),
            random.choice("~!@#$%&*")
        ]
        extra = random.choices(string.ascii_letters + string.digits + "~!@#$%&*", k=length - len(pools))
        pwd = pools + extra
        random.shuffle(pwd)
        s = "".join(pwd)
        return s if username.lower() not in s.lower() else self.generate_password(username, length)

    def rest_service(self, requests, method_name, params=None):
        if params is None:
            params = []
        params = [param if isinstance(param, str) else json.dumps(param) for param in params]
        payload = {
            "bindingName": "SR1450SRV1stuAppREST",
            "method": method_name,
            "params": params
        }
        print(f"Sending to rest service with medthod name: {method_name}")
        r = requests.post("https://ss2.sfcollege.edu/sr/restservices/SR1450SRV1stuAppREST", json=payload)
        print(r.text, r.status_code)
        
        if r.status_code != 200:
            return None
        try:
            # if r.json().get("result", {}).get("callReturnBR", {}).get("exceptionBL"):
            #     return None
            return r.json()
        except Exception:
            return None

    def close_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None
            
    def run(self, domain="tempmail.id.vn"):
        if not domain:
            return None
        session = requests.Session()
        user_name = self.generate_username()
        print(f"{user_name}@{domain}")
        
        r = session.post("https://ss2.sfcollege.edu/contact/information/check", json={"email": f"{user_name}@{domain}"})
        print(r.text)
        if r.status_code != 200:
            return None
        try:
            if r.json().get("response_code") == "ERROR":
                return None
        except Exception:
            return None
        
        email_info = self.client.create_temp_email(user_name, domain)
        if not email_info:
            return None
        email = email_info["email"]
        mail_id = email_info["id"]
        profile = self.fake_profile()
        print(email_info)
        print(profile)

        param1 = {
            "callReturnBR": {
                "exceptionBL":False,
                "referenceCdSTR":""
            },
            "stuNameSTR":"",
            "statusCitizenshipCdSTR":"U",
            "statusAdmissionCdSTR":"F",
            "statusHsCdSTR":"N",
            "statusCollegeExpCdSTR":"",
            "flPrepaidFlSTR":"",
            "dualEnrolFlSTR":"",
            "showHsCompFL":"",
            "showCollLvlFL":"",
            "showParentEduFL":"",
            "statusAdmissCdSTR":"",
            "statusStuInfoCdSTR":"",
            "statusContactCdSTR":"",
            "statusAddrPermCdSTR":"",
            "statusAddrParentCdSTR":"",
            "statusHighSchlCdSTR":"",
            "statusCollegeCdSTR":"",
            "statusCampusCdSTR":"",
            "statusDegreeCdSTR":"",
            "statusPrimLangCdSTR":"",
            "statusResidencyCdSTR":"",
            "statusDisciplinaryCdSTR":"",
            "statusParentEdCdSTR":"",
            "statusMilitaryStatCdSTR":"",
            "statusGenderRaceCdSTR":"",
            "isAchieve":False
        }
        param2 = {
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "fstNamSTR": profile["first_name"],
            "midNamSTR": "",
            "lstNamSTR": profile["last_name"],
            "sfxNamSTR": "",
            "prfNamSTR": "",
            "pronounIdINT": "3",
            "dobDtDT": profile["birthdate"].strftime("%m/%d/%Y"),
            "emailAltDesSTR": email,
            "birthCtryCdSTR": "045",
            "ssnumSTR": profile["ssn"].replace("-", ""),
            "alienNumSTR": "",
            "visaTypeCdSTR": "",
            "birthCitySTR": "",
            "stStuInfoCdSTR": "C",
            "citizenCtryCdSTR": "",
            "acceptPrivacyFlSTR": "",
            "primaryPhoneNumSTR": "",
            "nickNameSTR": profile["last_name"]
        }
        r = self.rest_service(session, "validateNewStudent", [param1, param2])
        if not r:
            return None

        verification_code_payload = {
            "id": None,
            "contact_type": "EMAIL",
            "contact_information": email,
            "contact_description": "Alt Email"
        }
        try:
            print("Verification code requesting")
            r = session.post("https://ss2.sfcollege.edu/contact/information/add", json=verification_code_payload)
            print(r.text, r.status_code)

            # if r and r.get("response_code") == "ERROR" and r.get("message") == "Duplicate Contact Information":
            #     r = session.post("https://ss2.sfcollege.edu/contact/information/token", json=verification_code_payload)
            #     print(r.text, r.status_code)

            id = r.json().get("payload", {}).get("id")
        except Exception:
            return None
        
        time.sleep(self.wait_sec)
        msg_id = self.client.get_message_by_match(mail_id, by_sender="no-reply@sfcollege.edu", by_subject="Application Code")
        if not msg_id:
            return None
        
        msg = self.client.read_message(msg_id)
        if not msg:
            return None
        code = self.extract_code(msg.get("body"))
        if not code:
            return None
        r = session.put("https://ss2.sfcollege.edu/contact/information/validate", json={"id": id, "validation_token": code})
        if r.status_code != 200:
            return None

        captchaToken = self.get_captcha_token()
        if not captchaToken:
            return None
        
        password = self.generate_password(profile["first_name"])
        resp = self.rest_service(session, "setAccountCreate", [
            param1,
            param2,
            {
                "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
                "psdSTR": password,
                "secQues1STR": "",
                "secQues2STR": "",
                "secQues3STR": ""
            },
            "",
            captchaToken
        ])
        if not resp:
            return None
        resp = session.post("https://ss2.sfcollege.edu/sr/stuappauth?api=getSession")
        print(resp.text)
        if resp.status_code != 200:
            return None
        try:
            student_id = resp.json().get("id")
        except Exception:
            return None
        print(f"{student_id} | {password}")

        # r = self.rest_service("getAccountInfo")
        # try:
        #     data = r.get("result")
        # except Exception:
        #     return None
        # r = self.rest_service("setMenuStatus", [data])
        # if not r:
        #     return None
        
        if not self.state_drop:
            r = self.rest_service(session, "getStateDrop")
        
            try:
                self.state_drop = r.get("result", {}).get("stateDropDown", [])
            except Exception:
                return None

        if not self.state_drop:
            return None

        state_code = next((item.get("stateCdSTR") for item in self.state_drop if item.get("stateNamSTR") == profile["state"]), None)
        if not state_code:
            return None
        print(state_code)
        r = self.rest_service(session, "setAddressCurrent", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "curAddress": {
                "countryCdSTR": "045",
                "street1NamSTR": profile["street"],
                "aptNumSTR": "",
                "cityNamSTR": profile["city"],
                "stateCdSTR": state_code,
                "countyCdSTR": "",
                "zip5CdSTR": profile["zip"],
                "provinceNamSTR": "",
                "postalNameSTR": ""
            },
            "phoneCellNumSTR": "",
            "curPhoneNumSTR": "",
            "emerPhonNumSTR": "",
            "cellTxtFlSTR": "Y",
            "statusContactCdSTR": "C"
        }])
        if not r:
            return None
        
        r = self.rest_service(session, "setAddressPermanent", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "permAddress": {
                "countryCdSTR": "",
                "street1NamSTR": "",
                "aptNumSTR": "",
                "cityNamSTR": "",
                "stateCdSTR": "",
                "countyCdSTR": "",
                "zip5CdSTR": "",
                "provinceNamSTR": "",
                "postalNameSTR": ""
            },
            "statusAddrPermCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setHighSchool", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "hsGradTypeCdSTR": "N",
            "hsSfccCdSTR": "",
            "hsSchNamSTR": "",
            "hsDiplomaNamSTR": "",
            "hsCountryNamSTR": "",
            "hsStateNamSTR": "",
            "hsCountyNamSTR": "",
            "dualEnrolFlSTR": "",
            "statusHighSchlCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setEnrollment", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "firstTermCdSTR": "2",
            "firstYrNumSTR": "2025",
            "appSiteCdSTR": "01",
            "finAidCdSTR": "Y",
            "statusCampusCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setDegreeSelection", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "curDegCdSTR": "ABE",
            "curProgCdSTR": "4100",
            "requestDegCdSTR": "",
            "requestProgCdSTR": "",
            "reqestedDegDesSTR": "",
            "curProgDesSTR": "",
            "statusDegreeCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setPrimaryLanguage", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "homeLangFL": "E",
            "homeLangNamSTR": "",
            "statusPrimLangCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setDisciplinary", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "disciplinarySchoolInd": "N/A",
            "disciplinaryViolenceInd": "No",
            "disciplinaryPredatorInd": "N/A",
            "statusDisciplinaryCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setStatsParentEd", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "highEdPg1CdSTR": "X",
            "highEdPg2CdSTR": "X",
            "statusParentEdCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setStatsMilitary", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "milStatCdSTR": "Y",
            "giBillFlSTR": "",
            "statusMilitaryStatCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "setStatsRace", [{
            "callReturnBR": {"exceptionBL": False, "referenceCdSTR": ""},
            "genderCdSTR": "M" if profile["gender"] == "Male" else "F",
            "ethnicHispanicFlSTR": "X",
            "raceWhiteFlSTR": "",
            "raceBlackFlSTR": "",
            "raceAsianFlSTR": "",
            "raceIndianSTR": "",
            "raceHawaiianFlSTR": "",
            "statusGenderRaceCdSTR": "C"
        }])
        if not r:
            return None
        r = self.rest_service(session, "activateApplication")
        if not r:
            return None
        
        session.post("https://ss2.sfcollege.edu/sr/stuappauth?api=clearSes")
        
        result = {
            "Email":        email,
            "Student ID":   student_id,
            "Password":     password,
            "Full Name":    f"{profile['first_name']} {profile['last_name']}",
            "Gender":       profile['gender'],
            "Birthdate":    profile['birthdate'].strftime("%Y-%m-%d"),
            "Street":       profile['street'],
            "City":         profile['city'],
            "State":        profile['state'],
            "Zipcode":      profile['zip'],
            "SSN":          profile['ssn'],
            "Password changed": "n"
        }
        return result

# def main():
#     data_file = "data.xlsx"
#     backup_file = "backup.xlsx"
#     headers = ["Email", "Student ID", "Password", "Full Name", "Gender", "Birthdate", "Street", "City", "State", "Zipcode", "SSN", "Password changed"]
#     try:
#         tries = max(int(input("Nhập số lượng (1): ")), 1)
#     except ValueError:
#         tries = 1

#     if os.path.exists(data_file):
#         workbook = load_workbook(data_file)
#         sheet = workbook.active

#         workbook.save(backup_file)
#         print(f"Đã sao lưu dữ liệu vào tệp: {backup_file}")
#     else:
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.title = "Data"
#         sheet.append(headers)

#     bot = Bot(token="934|f2ipqq9C8rbAvmSaIpp385XBpZTPkwuTGfgVMT4Sb826f99c")
#     success_count = 0

#     try:
#         for i in range(tries):
#             print(f"lần thử {i + 1}/{tries}")
#             result = bot.run()
#             if not result:
#                 print("Không thể tạo tài khoản")
#                 print(f"Tạo tài khoản thành công {success_count}/{tries}")
#                 continue

#             sheet.append([result[header] for header in headers])
#             workbook.save(data_file)
#             success_count += 1
#             print(f"Tạo tài khoản thành công {success_count}/{tries}")
#     except KeyboardInterrupt:
#         print("Chương trình đã dừng")
#     finally:
#         bot.close_driver()
# if __name__ == "__main__":
#     main()